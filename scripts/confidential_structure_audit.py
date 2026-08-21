"""Read-only, anonymised structural audit for confidential Excel references."""
from __future__ import annotations

import io
import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SAFE_LABELS = {
    "project", "project no", "project number", "project code", "description",
    "hours", "total", "totals", "name", "employee", "staff", "date", "month",
    "carried", "carryover", "status", "admin", "research", "training", "travel",
    "bids", "finance", "it", "team meetings", "holiday", "sick leave", "time in lieu",
}
MONTH_RE = re.compile(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s*[-/]?\s*\d{2,4}$", re.I)


def safe_label(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = " ".join(value.strip().lower().split()).rstrip(":")
    if normalized in SAFE_LABELS:
        return normalized
    if MONTH_RE.match(normalized):
        return "<month-label>"
    return None


def audit_workbook(source: object, index: int) -> dict:
    wb = load_workbook(source, read_only=False, data_only=False, keep_links=False)
    sheets = []
    for sheet_index, ws in enumerate(wb.worksheets):
        types = Counter()
        formulas = 0
        nonempty = 0
        safe_positions: list[str] = []
        date_heavy_rows: list[int] = []
        numeric_code_bands = Counter()
        for row in ws.iter_rows():
            dates_in_row = 0
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty += 1
                types[cell.data_type] += 1
                if cell.data_type == "f":
                    formulas += 1
                if getattr(cell, "is_date", False):
                    dates_in_row += 1
                label = safe_label(value)
                if label:
                    safe_positions.append(f"{cell.coordinate}:{label}")
                if isinstance(value, (int, float)) and float(value).is_integer():
                    number = int(value)
                    if 1000 <= number < 10000:
                        numeric_code_bands["1000-9999"] += 1
                    elif 10000 <= number < 100000:
                        numeric_code_bands["10000-99999"] += 1
            if dates_in_row >= 5:
                date_heavy_rows.append(row[0].row)
        sheets.append({
            "sheet_index": sheet_index,
            "visibility": ws.sheet_state,
            "rows": ws.max_row,
            "columns": ws.max_column,
            "nonempty_cells": nonempty,
            "cell_types": dict(types),
            "formula_count": formulas,
            "merged_range_count": len(ws.merged_cells.ranges),
            "date_heavy_rows": date_heavy_rows[:10],
            "safe_label_positions": safe_positions[:80],
            "numeric_code_bands": dict(numeric_code_bands),
        })
    return {"workbook_index": index, "sheet_count": len(sheets), "sheets": sheets}


def main(zip_path: str, monthly_path: str) -> None:
    with zipfile.ZipFile(zip_path) as archive:
        members = [item for item in archive.infolist() if not item.is_dir()]
        excel_members = [item for item in members if item.filename.lower().endswith((".xlsx", ".xlsm"))]
        workbooks = []
        failures = 0
        for index, member in enumerate(excel_members):
            try:
                workbooks.append(audit_workbook(io.BytesIO(archive.read(member)), index))
            except Exception:
                failures += 1
    result = {
        "timesheet_archive": {
            "file_count": len(members),
            "excel_file_count": len(excel_members),
            "non_excel_file_count": len(members) - len(excel_members),
            "unreadable_workbook_count": failures,
            "workbooks": workbooks,
        },
        "monthly_workbook": audit_workbook(Path(monthly_path), 0),
    }
    print(json.dumps(result, indent=2))


def value_kind(cell) -> str:
    value = cell.value
    if value is None:
        return "blank"
    if cell.data_type == "f":
        match = re.match(r"=\s*([A-Z][A-Z0-9.]*)", str(value), re.I)
        return f"formula:{(match.group(1) if match else 'expression').upper()}"
    label = safe_label(value)
    if label:
        return f"safe-label:{label}"
    if getattr(cell, "is_date", False):
        return "date"
    if isinstance(value, (int, float)):
        if 1000 <= value < 10000 and float(value).is_integer():
            return "numeric:project-code-band"
        if 10000 <= value < 100000 and float(value).is_integer():
            return "numeric:internal-code-band"
        return "numeric"
    if isinstance(value, str):
        return "text:suppressed"
    return cell.data_type


def compact_sheet(ws) -> dict:
    top = []
    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), max_col=min(40, ws.max_column)):
        for cell in row:
            kind = value_kind(cell)
            if kind != "blank":
                top.append(f"{cell.coordinate}:{kind}")
    columns = []
    for col in range(1, ws.max_column + 1):
        counts = Counter(value_kind(ws.cell(row, col)) for row in range(1, ws.max_row + 1))
        counts.pop("blank", None)
        columns.append({"column": get_column_letter(col), "nonblank_kinds": dict(counts)})
    repeated_text = []
    for row in range(1, min(10, ws.max_row) + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell.value, str) and not safe_label(cell.value):
                repeated_text.append(cell.coordinate)
    return {"rows": ws.max_row, "columns": ws.max_column, "top_cells": top, "column_profiles": columns, "suppressed_top_text_positions": repeated_text}


def compact_main(zip_path: str, monthly_path: str) -> None:
    with zipfile.ZipFile(zip_path) as archive:
        member = next(item for item in archive.infolist() if item.filename.lower().endswith((".xlsx", ".xlsm")))
        timesheet = load_workbook(io.BytesIO(archive.read(member)), read_only=False, data_only=False, keep_links=False)
    monthly = load_workbook(monthly_path, read_only=False, data_only=False, keep_links=False)
    result = {
        "timesheet_first_workbook": [compact_sheet(ws) for ws in timesheet.worksheets[:3]],
        "timesheet_sheet_dimensions": [[ws.max_row, ws.max_column] for ws in timesheet.worksheets],
        "monthly_workbook": [compact_sheet(ws) for ws in monthly.worksheets],
    }
    print(json.dumps(result, indent=2))


def fill_rgb(cell) -> str:
    fill = cell.fill
    if fill.fill_type != "solid":
        return "none" if fill.fill_type is None else f"pattern:{fill.fill_type}"
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:].upper()
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}"
    return "solid:unknown"


def acceptance_sheet(ws) -> dict:
    header_row = None
    for row in range(1, min(ws.max_row, 40) + 1):
        value = ws.cell(row, 2).value
        if isinstance(value, str) and value.strip().lower().startswith("job number"):
            header_row = row
            break
    result = {
        "sheet": ws.title if MONTH_RE.match(ws.title.strip()) else "<non-month-sheet>",
        "header_row": header_row,
        "rows": ws.max_row,
        "columns": ws.max_column,
    }
    if not header_row:
        return result
    carry_column = None
    notes_column = None
    for column in range(1, ws.max_column + 1):
        value = str(ws.cell(header_row, column).value or "").strip().lower()
        if "carried" in value or "carry" in value:
            carry_column = column
        if value == "notes":
            notes_column = column
    last_employee_column = (carry_column or notes_column or ws.max_column) - 1
    fill_counts = Counter()
    positive_numeric_cells = 0
    green_cells = 0
    green_rows = set()
    green_missing_project = 0
    for row in range(header_row + 2, ws.max_row + 1):
        for column in range(4, last_employee_column + 1):
            cell = ws.cell(row, column)
            if not isinstance(cell.value, (int, float)) or cell.value <= 0:
                continue
            positive_numeric_cells += 1
            fill = fill_rgb(cell)
            fill_counts[fill] += 1
            if fill == "92D050":
                green_cells += 1
                green_rows.add(row)
                if ws.cell(row, 2).value in (None, ""):
                    green_missing_project += 1
    legend = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, header_row)):
        for cell in row:
            text = str(cell.value or "").strip().lower()
            if "hrs to be invoiced" in text:
                legend["awaiting_invoice"] = fill_rgb(cell)
            elif "invoices sent" in text:
                legend["invoices_sent"] = fill_rgb(cell)
            elif "need to be carried" in text:
                legend["carry"] = fill_rgb(cell)
            elif "carried but have now been invoiced" in text:
                legend["carried_then_invoiced"] = fill_rgb(cell)
    result.update({
        "first_data_row": header_row + 2,
        "employee_column_count": max(0, last_employee_column - 3),
        "carry_column_present": carry_column is not None,
        "notes_column_present": notes_column is not None,
        "positive_numeric_employee_cells": positive_numeric_cells,
        "numeric_employee_fill_counts": dict(sorted(fill_counts.items())),
        "green_carry_cells": green_cells,
        "green_carry_row_count": len(green_rows),
        "green_carry_cells_missing_project_number": green_missing_project,
        "legend_fills": legend,
    })
    return result


def acceptance_main(monthly_path: str) -> None:
    workbook = load_workbook(monthly_path, read_only=False, data_only=False, keep_links=False)
    result = {
        "sheet_count": len(workbook.worksheets),
        "sheets": [acceptance_sheet(ws) for ws in workbook.worksheets],
    }
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    if len(sys.argv) > 3 and sys.argv[3] == "--acceptance":
        acceptance_main(sys.argv[2])
    elif len(sys.argv) > 3 and sys.argv[3] == "--compact":
        compact_main(sys.argv[1], sys.argv[2])
    else:
        main(sys.argv[1], sys.argv[2])
